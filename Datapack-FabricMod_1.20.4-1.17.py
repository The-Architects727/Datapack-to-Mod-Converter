import os
import shutil
import subprocess
import io
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog

original_directory = os.getcwd()

ModFilePath = None

failed_versions = []

complete = False

class InputGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Datapack to Fabric Mod Converter v. 1.20.4-1.17")
        self.master.geometry("430x500")

        # Create StringVar variables to store input values
        self.input_vars = [tk.StringVar() for _ in range(5)]
        self.folder_path_var = tk.StringVar()
        self.use_folder_var = tk.IntVar(value=0)  # Variable to track the selected option

        self.dest_path_var = tk.StringVar()



        # Custom labels for each input
        labels = ["Mod Name:", "Mod Description:", "Author:", "Datapack Github URL:"]

        # Create a button to open a folder selection dialog
        folder_button = tk.Button(self.master, text="Select Destination Folder", command=self.select_dest)
        folder_button.grid(row=10, column=0, pady=10, columnspan=2)

        #Create a label to display the selected folder path
        self.dest_label = tk.Label(self.master, textvariable=self.dest_path_var, wraplength=300)
        self.dest_label.grid(row=11, column=0, columnspan=2, pady=10)


        
        #labels = ["Mod Name:", "Mod Description:", "Author:"]

        # Create labels and entry widgets for each input
        for i in range(4):
            label_text = labels[i]
            label = tk.Label(self.master, text=label_text)
            location = i + 1
            if(i < 3):
                label.grid(row=location, column=0, padx=10, pady=10, sticky=tk.E)

                entry = tk.Entry(self.master, textvariable=self.input_vars[i])
                entry.grid(row=location, column=1, padx=10, pady=10, sticky=tk.W)
            else:
                label.grid(row=7, column=0, padx=10, pady=10, sticky=tk.E)

                entry = tk.Entry(self.master, textvariable=self.input_vars[i])
                entry.grid(row=7, column=1, padx=10, pady=10, sticky=tk.W)
            

        #Create a label to display the selected folder path
        self.folder_label = tk.Label(self.master, textvariable=self.folder_path_var, wraplength=300)
        self.folder_label.grid(row=9, column=0, columnspan=2, pady=10)

        
        # Create a title label
        title_label = tk.Label(self.master, text="Select Datapack Source:")
        title_label.grid(row=4, column=0, columnspan=2, pady=5)
        
        # Create a radio button for selecting the source of folder path
        folder_radio = tk.Radiobutton(self.master, text="Use Github URL", variable=self.use_folder_var, value=0)
        folder_radio.grid(row=5, column=0, padx=10, pady=5)
        folder_radio = tk.Radiobutton(self.master, text="Use Folder Select", variable=self.use_folder_var, value=1)
        folder_radio.grid(row=5, column=1, padx=10, pady=5)
        
        # Create a button to open a folder selection dialog
        folder_button = tk.Button(self.master, text="Select Local Datapack", command=self.select_folder)
        folder_button.grid(row=8, column=0, pady=10, columnspan=2)
        
        # Create a button to submit the inputs
        submit_button = tk.Button(self.master, text="Submit", command=self.submit_inputs)
        submit_button.grid(row=14, column=0, columnspan=2, pady=10)


    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path_var.set(folder_selected)
        self.folder_label.config(text=folder_selected)

    def select_dest(self):
        dest_selected = filedialog.askdirectory()
        self.dest_path_var.set(dest_selected)
        self.dest_label.config(text=dest_selected)

        
    def submit_inputs(self):
        is_url = False
        # Retrieve values from StringVar variables
        input_values = [var.get() for var in self.input_vars]

        # Determine the folder path based on the selected option
        if self.use_folder_var.get() == 0:  # If "Use String Input" is chosen
            folder_path = self.input_vars[3].get()  # Use the input for the last name as folder path
            is_url = True
        else:
            is_url = False
            folder_path = self.folder_path_var.get()

        # Print or use the input values as needed
        #print("Input values:", input_values)
        
        simple_mod_url = "https://github.com/The-Architects727/SimpleFabricModZip"  
        datapack_url = str(folder_path).replace("\\","/") 
        #datapack_url = "https://github.com/The-Architects727/BetterEndCitiesVanilla"#
        folder_from_repo2 = "data"  # Replace with the folder you want to copy from the second repository

        author = input_values[2]
        mod_name = input_values[0]
        mod_description = input_values[1]
        
        #author = 'TheArchitects'#
        #mod_name = 'Better End Cities Datapack Fabric Mod'#
        #mod_description = 'Better End Cities Datapack Made into a Mod for ease of use.'#
        
        dest_folder = os.path.join(self.dest_path_var.get(), remove(mod_name))
        modid = remove(mod_name).lower()
        if(dest_folder):
            if(mod_name):
                if(mod_description):
                    if(author):
                        if(datapack_url):
                            if(is_url):
                                self.master.destroy()
                                combine_repositories_url(mod_name, mod_description, modid, author, simple_mod_url, datapack_url, dest_folder, folder_from_repo2)
                            else:
                                self.master.destroy()
                                combine_repositories_folder(mod_name, mod_description, modid, author, simple_mod_url, datapack_url, dest_folder, folder_from_repo2)
                        else:
                            print('Datapack Location is Invalid')
                    else:
                        print('Author is Empty')
                else:
                    print('Mod Description is Empty')
            else:
                print('Mod Name is Empty')
        else:
            print('Destination Folder is Invalid')
        
                        
            
        #combine_repositories(mod_name, mod_description, modid, author, simple_mod_url, datapack_url, dest_folder, folder_from_repo2)
        #combine_repositories(simple_mod_url, datapack_url, dest_folder, folder_from_repo2)


def run_command_in_folder(folder_path, command):
    try:
        # Change directory to the specified folder
        os.chdir(folder_path)
        
        # Open command prompt and run the command using os.system()
        os.system(command)
    except Exception as e:
        print(f"Error: {e}")

def clone_repository(repo_url, destination):
    subprocess.run(['git', 'clone', repo_url, destination])

def copy_folder(src_folder, dest_folder):
    shutil.copytree(src_folder, dest_folder)

def replace_string_in_file(file_path, old_string, new_string):
    with open(file_path, 'r') as file:
        file_content = file.read()
    
    new_content = file_content.replace(old_string, new_string)
    
    with open(file_path, 'w') as file:
        file.write(new_content)

def replace_string_in_folder(folder_path, file_path, old_string, new_string):
    old_folder_path = os.getcwd()
    os.chdir(folder_path)
    
    #print('folder_path: ' + os.getcwd())
    
    #print('file_path: ' + file_path)
    #full_file_path = os.path.join(folder_path, file_path)
    replace_string_in_file(file_path, old_string, new_string)
    #full_file_path = os.path.join(folder_path, file_path, "")
    #replace_string_in_file(full_file_path, old_string, new_string)

    os.chdir(old_folder_path)
    
    #print('folder_path: ' + os.getcwd())

def print_file_contents(folder_path, file_path):
    old_folder_path = os.getcwd()
    os.chdir(folder_path)
    try:
        with open(file_path, 'r') as file:
            contents = file.read()
            print(contents)
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")
    os.chdir(old_folder_path)

def combine_repositories_url(mod_name, mod_description, modid, author, repo1_url, repo2_url, dest_folder, folder_from_repo2):
    # Clone the first repository
    repo1_folder = "repo1"
    clone_repository(repo1_url, repo1_folder)

    # Clone the second repository
    repo2_folder = "repo2"
    clone_repository(repo2_url, repo2_folder)

    # Copy the entire contents of repo1 to the destination folder
    copy_folder(repo1_folder, dest_folder)

    # Specify the subfolder within repo1 to copy the folder from repo2 into
    subfolder_within_repo1 = "src/main/resources/data"  # Replace with the desired subfolder name

    # Copy only the specified folder from repo2 to the specified subfolder in repo1
    src_folder_from_repo2 = os.path.join(repo2_folder, folder_from_repo2)
    dest_subfolder = os.path.join(dest_folder, subfolder_within_repo1)
    copy_folder(src_folder_from_repo2, dest_subfolder)

    # Specify the file path and strings to replace(Mod Name)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'Example mod'
    new_string = mod_name

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    #print('changing name: ' + os.getcwd())

    # Specify the file path and strings to replace(Mod Description)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'This is an example description! Tell everyone what your mod is about!'
    new_string = mod_description

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    # Specify the file path and strings to replace(Author)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'Me!'
    new_string = author

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    # Specify the file path and strings to replace(Mod Description)
    file_to_modify = 'gradle.properties'
    old_string = 'simplemod'
    new_string = modid

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    replace_values(dest_folder, modid, mod_name)

    # Clean up: remove temporary repositories
    #shutil.rmtree(repo1)
    #shutil.rmtree(repo2)
    
    
def combine_repositories_folder(mod_name, mod_description, modid, author, repo1_url, repo2_folder, dest_folder, folder_from_repo2):
    # Clone the first repository
    repo1_folder = "repo1"
    clone_repository(repo1_url, repo1_folder)

    # Copy the entire contents of repo1 to the destination folder
    copy_folder(repo1_folder, dest_folder)

    # Specify the subfolder within repo1 to copy the folder from repo2 into
    subfolder_within_repo1 = "src/main/resources/data"  # Replace with the desired subfolder name

    # Copy only the specified folder from repo2 to the specified subfolder in repo1
    src_folder_from_repo2 = os.path.join(repo2_folder, folder_from_repo2)
    dest_subfolder = os.path.join(dest_folder, subfolder_within_repo1)
    copy_folder(src_folder_from_repo2, dest_subfolder)

    # Specify the file path and strings to replace(Mod Name)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'Example mod'
    new_string = mod_name

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    #print('changing name: ' + os.getcwd())

    # Specify the file path and strings to replace(Mod Description)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'This is an example description! Tell everyone what your mod is about!'
    new_string = mod_description

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    # Specify the file path and strings to replace(Author)
    file_to_modify = 'src/main/resources/fabric.mod.json'
    old_string = 'Me!'
    new_string = author

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    # Specify the file path and strings to replace(Mod Description)
    file_to_modify = 'gradle.properties'
    old_string = 'simplemod'
    new_string = modid

    # Replace string in the specified file
    replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

    replace_values(dest_folder, modid, mod_name)

    # Clean up: remove temporary repositories
    #shutil.rmtree(repo1)
    #shutil.rmtree(repo2)
    



def build_jar(dest_folder, minecraft_version, modid):
    #________BUILD_JAR________#
    time.sleep(5)
    # Example of running a command in the target folder
    command_to_run = 'gradlew build'
    run_command_in_folder(dest_folder, command_to_run)

    continue_wait = True

    old_folder_path = os.getcwd()

    counter = 0
    
    while (continue_wait):
        counter += 1
        #print(counter)
        try:
            folder_path = old_folder_path + "/build/libs"
            folder_path_new = folder_path.replace("\\","/")
            
            os.chdir(folder_path_new)
            
            #folder_path = (dest_folder + "/build/libs")
            file_name = (modid + "-" + minecraft_version + ".jar")

            #file_path = os.path.join(folder_path, file_name)

            if os.path.isfile(os.getcwd() + "/" + file_name):
                print(f"Completed for " + minecraft_version)
                continue_wait = False
            else:
                if (counter >= 5):
                    continue_wait = False
                    failed_versions.append(minecraft_version)
                time.sleep(5)
        except Exception as e:
            print(f"An error occurred: {e}")
            os.chdir(old_folder_path)
            time.sleep(5)
    
            
    os.chdir(old_folder_path)


def replace_current_values(modid, dest_folder, minecraft_version, fabric_version, mapping_version, loader_version, prior_minecraft_version, prior_fabric_version, prior_mapping_version, prior_loader_version):
    os.chdir(original_directory)
    #print('changing version: ' + os.getcwd())

    # home_path = 'D:/Main/Modding/fabric-example-mod-1.20-barebones/PYTHON/WithExcel/DatapackFabricMod'
    
    print("Starting: " + str(minecraft_version))
    if minecraft_version == '1.20.4':
        # Specify the file path and strings to replace(Minecraft Version)
        file_to_modify = r'gradle.properties'
        old_string = '1.0.0'
        new_string = minecraft_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

        #print_file_contents(dest_folder, file_to_modify)
        
        #Build Jar
        build_jar(dest_folder, minecraft_version, modid)
    else:
        #_______CHANGE VERSION_________
        
        # Specify the file path and strings to replace(Mod Description)
        #print('first change: ' + os.getcwd())
        file_to_modify = r'src/main/resources/fabric.mod.json'
        old_string = prior_minecraft_version
        new_string = minecraft_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)


        # Specify the file path and strings to replace(Mod Description)
        file_to_modify = r'gradle.properties'
        old_string = prior_fabric_version
        new_string = fabric_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)


        # Specify the file path and strings to replace(Mod Description)
        file_to_modify = r'gradle.properties'
        old_string = prior_mapping_version
        new_string = mapping_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)


        # Specify the file path and strings to replace(Mod Description)
        file_to_modify = r'gradle.properties'
        old_string = prior_loader_version
        new_string = loader_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)
        

        # Specify the file path and strings to replace(Minecraft Version)
        file_to_modify = r'gradle.properties'
        old_string = prior_minecraft_version
        new_string = minecraft_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)
        

        # Specify the file path and strings to replace(Minecraft Version)
        file_to_modify = r'gradle.properties'
        old_string = prior_minecraft_version
        new_string = minecraft_version

        # Replace string in the specified file
        replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

        #print_file_contents(dest_folder, file_to_modify)

        if(minecraft_version == '1.19.1'):
            # Specify the file path and strings to replace(Fabric api Fix)
            file_to_modify = r'src/main/resources/fabric.mod.json'
            old_string = 'fabric-api'
            new_string = 'fabric'

            # Replace string in the specified file
            replace_string_in_folder(dest_folder, file_to_modify, old_string, new_string)

        else:
            time.sleep(1)
        #Build Jar
        build_jar(dest_folder, minecraft_version, modid)

    


    
    
def replace_values(dest_folder, modid, mod_name):
    
    # Load the Excel workbook
    workbook = openpyxl.load_workbook('versions.xlsx')

    # Select the active sheet
    sheet = workbook.active

    
    #os.chdir(dest_folder)

    # Initialize variables for current and prior values
    prior_minecraft_version = '1.20.4'
    prior_fabric_version = '0.91.1+1.20.4'
    prior_mapping_version = '1.20.4+build.1'
    prior_loader_version = '0.15.0'

    max_repetitions = 15  # Set your desired maximum number of repetitions

    counter = 0  # Initialize a counter variable

    
    # Loop through the values in the first column (assuming values are in column A)
    for row in sheet.iter_rows(min_row=1, max_col=5, max_row=sheet.max_row, values_only=True):
        
        minecraft_version = str(row[0])
        fabric_version = str(row[2])
        mapping_version = str(row[3])
        loader_version = str(row[4])
        if(minecraft_version != None):

            #print(minecraft_version)
            #print(fabric_version)
            #print(mapping_version)
            #print(loader_version)
        
            #print(os.getcwd())
        
            replace_current_values(modid, dest_folder, minecraft_version, fabric_version, mapping_version, loader_version, prior_minecraft_version, prior_fabric_version, prior_mapping_version, prior_loader_version)
        
            prior_minecraft_version = minecraft_version
            prior_fabric_version = fabric_version
            prior_mapping_version = mapping_version
            prior_loader_version = loader_version
        else:
            time.sleep(1)

        # Increment the counter
        counter += 1
    
        # Check if the counter has reached the maximum repetitions
        if counter >= max_repetitions:
            break  # Exit the loop if the maximum repetitions are reached

    # Create a new window
    new_window = tk.Tk()
    new_window.title("Completed")

    # Add two titles to the new window
    title1_label = tk.Label(new_window, text=("Conversion Complete for " + mod_name))
    title1_label.grid(row=0, column=0, pady=10)

    title2_label = tk.Label(new_window, text=("Failed Versions: " + str(failed_versions)))
    title2_label.grid(row=1, column=0, pady=10)
        
    print("Conversion Complete")
    print("Failed Versions: " + str(failed_versions))
    # Close the workbook when done
    workbook.close()

    
def remove(string):
    return string.replace(" ", "")


if __name__ == "__main__":

    root = tk.Tk()
    app = InputGUI(root)
    root.mainloop()
    
    #simple_mod_url = "https://github.com/The-Architects727/SimpleFabricModZip"  
    #datapack_url = "https://github.com/The-Architects727/BetterEndCitiesVanilla"#
    #folder_from_repo2 = "data"  # Replace with the folder you want to copy from the second repository
    #author = 'TheArchitects'#
    #mod_name = 'Better End Cities Datapack Fabric Mod'#
    #mod_description = 'Better End Cities Datapack Made into a Mod for ease of use.'#
    #dest_folder = remove(mod_name)
   # modid = dest_folder.lower()

    #combine_repositories(mod_name, mod_description, modid, author, simple_mod_url, datapack_url, dest_folder, folder_from_repo2)
    #combine_repositories(simple_mod_url, datapack_url, dest_folder, folder_from_repo2)
